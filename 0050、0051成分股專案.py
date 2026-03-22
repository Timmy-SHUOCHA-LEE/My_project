# 載入所需的模組
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook, Workbook
import time

# 動態參數設置
chrome_driver_path = "C:\\python training\\股票專案\\chromedriver.exe"  # ChromeDriver 路徑
excel_file_path = "C:\\Users\\asus\\Desktop\\0421ETF標的 (1).xlsx"  # Excel 文件路徑
target_urls = [
    "https://www.yuantaetfs.com/product/detail/0050/ratio",
    "https://www.yuantaetfs.com/product/detail/0051/ratio"
]  # 需要提取數據的網站網址

# 初始化 WebDriver
options = Options()
options.chrome_executable_path = chrome_driver_path
driver = webdriver.Chrome(options=options)

# 定義整合後數據的容器
final_data = {"商品代碼": [], "商品名稱": []}

# 定義提取資料的函數
def extract_data(url, final_data):
    driver.get(url)
    time.sleep(2)

    try:
        # 點擊展開按鈕
        more_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, "moreBtn"))
        )
        more_button.click()
    except Exception as e:
        print("無法點擊展開按鈕:", e)

    try:
        titleTags = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, "td"))
        )
        data = [titleTag.text.strip() for titleTag in titleTags]

        temp_data = {"商品代碼": [], "商品名稱": []}
        for i in range(0, len(data) - 3, 4):
            if all("期貨" not in data[j].lower() for j in range(i, i + 4)):
                temp_data["商品代碼"].append(data[i].strip())
                temp_data["商品名稱"].append(data[i + 1].strip())
        temp_data["商品代碼"] = temp_data["商品代碼"][1:-1]
        temp_data["商品名稱"] = temp_data["商品名稱"][1:-1]
        final_data["商品代碼"].extend(temp_data["商品代碼"])
        final_data["商品名稱"].extend(temp_data["商品名稱"])
    except Exception as e:
        print("資料提取失敗:", e)

# 從多個網址提取數據
for url in target_urls:
    extract_data(url, final_data)

# 確保數據的一致性
final_data["商品代碼"] = [str(code).strip().lower() for code in final_data["商品代碼"]]
final_data["商品名稱"] = [str(name).strip() for name in final_data["商品名稱"]]

# 載入 Excel 檔案
workbook = load_workbook(filename=excel_file_path)
sheet = workbook.active

# 擷取 G 欄的代號並進行預處理
column_g_values = [str(cell.value).strip().lower() for cell in sheet['G'][1:] if cell.value is not None]

# 擷取 H 欄的一般個股標的名稱
column_h_values = [str(cell.value).strip() for cell in sheet['H'][1:] if cell.value is not None]

# 整合輸出資料
output_data = []

# 任務二：找出商品代碼有但 G 欄沒有的代號
differences = set(final_data["商品代碼"]) - set(column_g_values)
print("商品代碼有，但 G 欄沒有的代號：")
for difference in differences:
    index = final_data["商品代碼"].index(difference)
    action = "新增"
    etf = "0050" if "0050" in target_urls[0] else "0051"
    print(f"{difference} {final_data['商品名稱'][index]} : {action}")
    output_data.append([difference, final_data['商品名稱'][index], action, etf])

# 任務三：找出商品代碼沒有，但 G 欄有的代號
false_codes = set(column_g_values) - set(final_data["商品代碼"])
print("商品代碼沒有，但 G 欄有的代號：")
for code in false_codes:
    name = column_h_values[column_g_values.index(code)] if code in column_g_values else "未找到對應名稱"
    action = "剔除"
    etf = "未判定"
    print(f"{code} {name} {action}")
    output_data.append([code, name, action, etf])

# 關閉瀏覽器
driver.close()

# 輸出整合後的資料到新的 Excel 文件
new_workbook = Workbook()
new_sheet = new_workbook.active

# 設置標題行
new_sheet.append(["股票代號", "股票名稱", "新增或剔除", "ETF"])

# 將整合後的資料逐行寫入 Excel
for row in output_data:
    new_sheet.append(row)

# 保存新 Excel 文件
new_excel_file_path = "C:\\python training\\股票專案\\比對結果\\比對結果 (1).xlsx"  # 自訂路徑及檔名
new_workbook.save(new_excel_file_path)

print(f"已成功生成 Excel 檔案，存於: {new_excel_file_path}")
